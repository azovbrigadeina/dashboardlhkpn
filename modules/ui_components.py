import streamlit as st

def inject_custom_css():
    st.markdown("""
    <style>


    .metric-card {
        background-color: white;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.07);
        text-align: center;
        border-top: 5px solid #ececec;
    }
    .metric-label { font-size: 14px; color: #64748b; font-weight: bold; }
    .metric-value { font-size: 32px; font-weight: bold; color: #1e293b; margin: 5px 0; }
    .metric-delta { font-size: 13px; font-weight: 600; }

    .sync-log {
        background: #0f172a;
        color: #22d3ee;
        font-family: 'Courier New', monospace;
        font-size: 13px;
        border-radius: 10px;
        padding: 20px;
        line-height: 2;
        min-height: 200px;
    }
    .sync-ok  { color: #4ade80; }
    .sync-err { color: #f87171; }
    .sync-warn{ color: #fbbf24; }
    .sync-info{ color: #22d3ee; }
    </style>
    """, unsafe_allow_html=True)

def render_metric_card(label, value, delta_text, delta_color, border_color):
    st.markdown(f"""
    <div class="metric-card" style="border-top-color:{border_color};">
        <div class="metric-label">{label}</div>
        <div class="metric-value">{value}</div>
        <div class="metric-delta" style="color:{delta_color};">{delta_text}</div>
    </div>
    """, unsafe_allow_html=True)

def render_footer():
    st.write("---")
    st.markdown("""
    <div style="text-align:center; color:#94a3b8; font-size:12px; padding: 10px;">
        🏛️ LHKPN Monitoring System — Universitas Jambi &nbsp;|&nbsp; 
        Data bersumber from e-LHKPN KPK &nbsp;|&nbsp; 
        Sistem ini bersifat internal dan rahasia
    </div>
    """, unsafe_allow_html=True)

def render_executive_panel(data, paripurna_txt, count_u100, atensi_label, potential_pct, periode=""):
    st.markdown(f"""
    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:20px;">
        <div style="display:flex;gap:15px;flex-wrap:wrap;">
            <div style="flex:1;min-width:250px;background:white;padding:15px;border-radius:8px;border:1px solid #bbf7d0;">
                <b style="color:#166534;">🏆 APRESIASI</b><br>
                <small>Unit Paripurna: {paripurna_txt} ({count_u100} Unit)</small>
            </div>
            <div style="flex:1;min-width:250px;background:white;padding:15px;border-radius:8px;border:1px solid #fecaca;">
                <b style="color:#9f1239;">⚠️ ATENSI</b><br>
                <small>Prioritas: {atensi_label}</small>
            </div>
            <div style="flex:1;min-width:250px;background:white;padding:15px;border-radius:8px;border:1px solid #bfdbfe;">
                <b style="color:#1e40af;">⚡ AKSELERASI</b><br>
                <small>Potensi Maksimal: {potential_pct:.1f}% jika Zona Kuning tuntas.</small>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Add download button for executive report (admin only)
    if st.session_state.get('role') == 'admin' and data is not None:
        st.write("")
        from datetime import datetime
        periode_safe = periode.replace(" ", "_") if periode else "Akumulasi"
        fname = f"Laporan_Eksekutif_LHKPN_{periode_safe}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="📥 Unduh Laporan Eksekutif",
            data=generate_executive_report(data, periode),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_executive_report"
        )

def generate_executive_report(data, periode=""):
    """Generate a professionally formatted Excel executive report.
    Returns a BytesIO object for Streamlit's download_button.
    """
    import pandas as pd, io
    from datetime import datetime
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()

    # --- Compute metrics ---
    total_wl = len(data)
    h = len(data[data['ZONA'] == "🟢 ZONA HIJAU"])
    k = len(data[data['ZONA'] == "🟡 ZONA KUNING"])
    m = len(data[data['ZONA'] == "🔴 ZONA MERAH"])
    dl = len(data[data['Status LHKPN'].astype(str).str.strip() == "Diumumkan Lengkap"])
    rate = (h / total_wl * 100) if total_wl > 0 else 0
    dl_rate = (dl / total_wl * 100) if total_wl > 0 else 0

    # --- Style constants ---
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1E3A5F")
    subtitle_font = Font(bold=True, size=11, color="4A5568")
    value_font = Font(bold=True, size=12)
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    def auto_width(ws, min_w=12, max_w=40):
        for col_cells in ws.columns:
            length = max((len(str(c.value or "")) for c in col_cells), default=0)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(min_w, min(length + 4, max_w))

    def style_header_row(ws, row=1, cols=1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ========== SHEET 1: RINGKASAN EKSEKUTIF ==========
        ws_sum = writer.book.create_sheet("Ringkasan Eksekutif")
        writer.book.active = writer.book.sheetnames.index("Ringkasan Eksekutif")

        # Title block
        ws_sum.merge_cells('A1:D1')
        ws_sum['A1'].value = "LAPORAN EKSEKUTIF KEPATUHAN LHKPN"
        ws_sum['A1'].font = title_font
        ws_sum['A1'].alignment = Alignment(horizontal='center')

        ws_sum.merge_cells('A2:D2')
        ws_sum['A2'].value = "Universitas Jambi"
        ws_sum['A2'].font = subtitle_font
        ws_sum['A2'].alignment = Alignment(horizontal='center')

        ws_sum.merge_cells('A3:D3')
        ws_sum['A3'].value = f"Periode: {periode if periode else 'Global (Akumulasi)'}"
        ws_sum['A3'].font = Font(size=10, color="64748B")
        ws_sum['A3'].alignment = Alignment(horizontal='center')

        ws_sum.merge_cells('A4:D4')
        ws_sum['A4'].value = f"Dicetak: {datetime.now().strftime('%d %B %Y, %H:%M WIB')}"
        ws_sum['A4'].font = Font(size=9, color="94A3B8", italic=True)
        ws_sum['A4'].alignment = Alignment(horizontal='center')

        # Summary table starting at row 6
        sum_headers = ["Indikator", "Jumlah", "Persentase", "Keterangan"]
        for c, hdr in enumerate(sum_headers, 1):
            cell = ws_sum.cell(row=6, column=c, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        rows_data = [
            ("Total Wajib Lapor", total_wl, "100%", "Seluruh pegawai yang wajib melaporkan LHKPN", None),
            ("Zona Hijau (Sudah Lapor)", h, f"{rate:.1f}%", "Telah menyampaikan LHKPN", green_fill),
            ("Zona Kuning (Draft)", k, f"{(k/total_wl*100):.1f}%" if total_wl else "0%", "Draft tersimpan, belum disubmit", yellow_fill),
            ("Zona Merah (Belum Lapor)", m, f"{(m/total_wl*100):.1f}%" if total_wl else "0%", "Belum ada aktivitas pelaporan", red_fill),
            ("Diumumkan Lengkap (Paripurna)", dl, f"{dl_rate:.1f}%", "Telah diumumkan secara lengkap kepada publik", None),
        ]
        for i, (label, count, pct, desc, fill) in enumerate(rows_data):
            r = 7 + i
            ws_sum.cell(row=r, column=1, value=label).border = thin_border
            ws_sum.cell(row=r, column=2, value=count).border = thin_border
            ws_sum.cell(row=r, column=2).alignment = Alignment(horizontal='center')
            ws_sum.cell(row=r, column=2).font = value_font
            ws_sum.cell(row=r, column=3, value=pct).border = thin_border
            ws_sum.cell(row=r, column=3).alignment = Alignment(horizontal='center')
            ws_sum.cell(row=r, column=4, value=desc).border = thin_border
            if fill:
                for c in range(1, 5):
                    ws_sum.cell(row=r, column=c).fill = fill

        # Potensi akselerasi
        pot = ((h + k) / total_wl * 100) if total_wl > 0 else 0
        r_pot = 13
        ws_sum.merge_cells(f'A{r_pot}:D{r_pot}')
        ws_sum[f'A{r_pot}'].value = f"Potensi Kepatuhan Maksimal (jika seluruh Draft tuntas): {pot:.1f}%"
        ws_sum[f'A{r_pot}'].font = Font(bold=True, color="1E40AF", size=10)

        auto_width(ws_sum)
        ws_sum.column_dimensions['A'].width = 36
        ws_sum.column_dimensions['D'].width = 48

        # ========== SHEET 2: STATISTIK PER UNIT ==========
        unit_stats = data.groupby('SUB UNIT KERJA')['ZONA'].value_counts().unstack(fill_value=0)
        for z in ["🟢 ZONA HIJAU", "🟡 ZONA KUNING", "🔴 ZONA MERAH"]:
            if z not in unit_stats.columns:
                unit_stats[z] = 0
        unit_stats = unit_stats[["🟢 ZONA HIJAU", "🟡 ZONA KUNING", "🔴 ZONA MERAH"]]
        unit_stats['Total'] = unit_stats.sum(axis=1)
        unit_stats['% Kepatuhan'] = (unit_stats['🟢 ZONA HIJAU'] / unit_stats['Total'] * 100).round(1)

        # Count DL per unit
        dl_per_unit = data[data['Status LHKPN'].astype(str).str.strip() == "Diumumkan Lengkap"].groupby('SUB UNIT KERJA').size()
        unit_stats['Diumumkan Lengkap'] = dl_per_unit.reindex(unit_stats.index, fill_value=0).astype(int)
        unit_stats['% Paripurna'] = (unit_stats['Diumumkan Lengkap'] / unit_stats['Total'] * 100).round(1)

        # Rename columns for clean display
        unit_export = unit_stats.rename(columns={
            "🟢 ZONA HIJAU": "Hijau",
            "🟡 ZONA KUNING": "Kuning",
            "🔴 ZONA MERAH": "Merah",
        }).reset_index().rename(columns={"SUB UNIT KERJA": "Unit Kerja"})
        unit_export = unit_export.sort_values('% Kepatuhan', ascending=False)

        unit_export.to_excel(writer, sheet_name="Statistik Per Unit", index=False, startrow=1)
        ws_unit = writer.sheets["Statistik Per Unit"]
        ws_unit.merge_cells('A1:H1')
        ws_unit['A1'].value = f"Statistik Kepatuhan LHKPN Per Unit Kerja — {periode if periode else 'Global'}"
        ws_unit['A1'].font = subtitle_font

        ncols = len(unit_export.columns)
        style_header_row(ws_unit, row=2, cols=ncols)

        # Style data rows
        for r in range(3, len(unit_export) + 3):
            for c in range(1, ncols + 1):
                cell = ws_unit.cell(row=r, column=c)
                cell.border = thin_border
                if c >= 2:
                    cell.alignment = Alignment(horizontal='center')

        # Total row
        total_row = len(unit_export) + 3
        ws_unit.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for c_idx, col_name in enumerate(unit_export.columns, 1):
            cell = ws_unit.cell(row=total_row, column=c_idx)
            cell.border = thin_border
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            if col_name in ['Hijau', 'Kuning', 'Merah', 'Total', 'Diumumkan Lengkap']:
                cell.value = int(unit_export[col_name].sum())
            elif col_name == '% Kepatuhan':
                cell.value = f"{rate:.1f}%"
            elif col_name == '% Paripurna':
                cell.value = f"{dl_rate:.1f}%"

        auto_width(ws_unit)
        ws_unit.column_dimensions['A'].width = 35

        # ========== SHEET 3: DATA DETAIL ==========
        # Clean columns — remove internal helper columns and the original "No." column to prevent duplicates
        detail_cols = [c for c in data.columns if c not in ['rank', 'NIK_KEY', 'No.']]
        detail_data = data[detail_cols].copy()

        # Sort by Status LHKPN: Terverifikasi Lengkap first, then others in a logical sequence
        status_order = {
            'Terverifikasi Lengkap': 1,
            'Diumumkan Lengkap': 2,
            'Proses Verifikasi': 3,
            'Perlu Perbaikan': 4,
            'Diumumkan Tidak Lengkap': 5,
            'Draft': 6,
            'Belum Lapor': 7
        }
        detail_data['status_sort_idx'] = detail_data['Status LHKPN'].astype(str).str.strip().map(status_order).fillna(99)
        detail_data = detail_data.sort_values('status_sort_idx').drop(columns=['status_sort_idx']).reset_index(drop=True)

        # Add sequential No column at the front
        detail_data.insert(0, 'No', range(1, len(detail_data) + 1))

        # Force NIK as string to preserve leading zeros
        if 'NIK' in detail_data.columns:
            detail_data['NIK'] = detail_data['NIK'].astype(str).str.strip()

        # Remove emoji from ZONA for cleaner export
        detail_data['ZONA'] = detail_data['ZONA'].str.replace(r'[🟢🟡🔴⚪]\s*', '', regex=True)

        all_detail_cols = list(detail_data.columns)
        detail_data.to_excel(writer, sheet_name="Data Detail", index=False, startrow=1)
        ws_detail = writer.sheets["Data Detail"]

        ncols_d = len(all_detail_cols)
        ws_detail.merge_cells(f'A1:{get_column_letter(ncols_d)}1')
        ws_detail['A1'].value = f"Data Detail Wajib Lapor LHKPN — {periode if periode else 'Global'}"
        ws_detail['A1'].font = subtitle_font

        style_header_row(ws_detail, row=2, cols=ncols_d)

        # Find column indices for special handling
        zona_col_idx = all_detail_cols.index('ZONA') + 1 if 'ZONA' in all_detail_cols else None
        nik_col_idx = all_detail_cols.index('NIK') + 1 if 'NIK' in all_detail_cols else None
        wrap_align = Alignment(vertical='center', wrap_text=True)
        center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Style data rows + conditional zona coloring + wrap text
        for r in range(3, len(detail_data) + 3):
            for c in range(1, ncols_d + 1):
                cell = ws_detail.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = wrap_align
                # Force NIK as text so Excel doesn't mangle it
                if c == nik_col_idx:
                    cell.number_format = '@'
                    cell.value = str(cell.value) if cell.value is not None else ""
                # Center the No column
                if c == 1:
                    cell.alignment = center_wrap
            if zona_col_idx:
                zona_val = str(ws_detail.cell(row=r, column=zona_col_idx).value or "")
                row_fill = None
                if "HIJAU" in zona_val:
                    row_fill = green_fill
                elif "KUNING" in zona_val:
                    row_fill = yellow_fill
                elif "MERAH" in zona_val:
                    row_fill = red_fill
                if row_fill:
                    for c in range(1, ncols_d + 1):
                        ws_detail.cell(row=r, column=c).fill = row_fill

        # Set sensible column widths (not too wide)
        for col_cells in ws_detail.columns:
            col_letter = get_column_letter(col_cells[0].column)
            col_idx = col_cells[0].column
            header_name = all_detail_cols[col_idx - 1] if col_idx <= len(all_detail_cols) else ""
            if header_name == 'No':
                ws_detail.column_dimensions[col_letter].width = 5
            elif header_name == 'NIK':
                ws_detail.column_dimensions[col_letter].width = 22
            elif header_name == 'NAMA':
                ws_detail.column_dimensions[col_letter].width = 28
            elif header_name == 'SUB UNIT KERJA':
                ws_detail.column_dimensions[col_letter].width = 30
            elif header_name in ('Status LHKPN', 'ZONA'):
                ws_detail.column_dimensions[col_letter].width = 22
            else:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
                ws_detail.column_dimensions[col_letter].width = max(8, min(max_len + 2, 30))

        # Remove default empty 'Sheet' if created
        if "Sheet" in writer.book.sheetnames:
            del writer.book["Sheet"]

    output.seek(0)
    return output
